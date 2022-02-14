"""Utility for load [xls|xlsx|csv|html] to oracle table v.0.0.3

Usage: xls2ora.exe file.ext|file.json

file.ext [xls|xlsx|csv|html]

Example config file:
* - required fields

xls2ora.json =>
{
*"table_in":"shtat.shtat_reports",
*"fields_in":"pib,time_inout,rdate",
"file_in":"file.ext", * if arg file.json
"first_row":2,
"cols":[2,3,"&filename"],
"format":"html|xls|csv",
"separator":",",
"truncate":"Y|n",
"delete":" rdate=\"&filename\""
}        
"""

import os
# from re import T
import sys
import time
import xlrd
import openpyxl
import bs4
import requests
from os import path
# from sys import exit
import json
import traceback
from funks import (
    file2arr,
    sendicqmsg,
    nm,
    decl_log,
    myLog,
    username,
    local_ip,
)

url_api = "http://10.9.19.15:5000/api"
headers = {"Content-Type": "application/json; charset=ISO-8859-1"}
proxies = {"http": "", "https": ""}


def truncate_table(table_in="",delete=""):
    um_delete=""
    if delete:
        um_delete=f" where {delete}".replace('"','\"')
    json={"action":"sql","sql":f"delete from {table_in} {um_delete}"}
    return request_api(json)

def ins_to_ora(data_in=[],table_in='',fields_in=''):
    if not data_in or not table_in or not fields_in:
        return "empty input data"
    fields_in=[f.lower().strip() for f in fields_in.split(",")]
    json_data = {
        "action": "executemany",
        "table": table_in,
        "fields":fields_in,
        "data":data_in,}
    
    return request_api(json_data)
    # url = f"""http://127.0.0.1:5000/api"""

def request_api(json_data):
    
    json_data["src"]= nm
    json_data["username"]= username
    json_data["user_ip"]= local_ip


    try:
        resp=requests.post(url_api,json=json_data, headers=headers,proxies=proxies)
    except Exception as e:
        myLog(f"error request api: {e}",1)
        return -1

    myLog(f"resp.status: {resp.status_code}")
    if resp.status_code != 200:
        myLog(f"error resp.status: {resp.status_code}",1)
        return -1
    else:
        try:
            data = resp.json()
            if data["cnt"]<0:
                myLog(f"""error exec sql: {data['result'][0][0]}""",1)
                return -1
            else:
                myLog(f"sql Ok: {data['cnt']} rows",1)
                return 1
        except Exception as e:
            myLog(f"""error get json: {e}""",1)
            return -1


def do_xls2ora():
    try:
        myLog("BEGIN")
        usage=__doc__
        try:    
            arg = sys.argv[1:][0]
        except:
            print(usage)
            return
            
        if arg in ['/?','--help','-h']:
            print(usage)
            return
        
        myLog(f"Utility for load [xls|xlsx|csv|html] to oracle table\n")
        # usage = "xls2ora.exe {filename} формату [XLS]\n!!!"

        if not path.exists(arg):
            myLog(f"""Input file [{arg}] not exists\n""",1)
            print(usage)
            sendicqmsg(1001,f"""Input file [{arg}] not exists""")
            return

        

        if arg.find(".json")>-1:
            json_file=arg
        else:
            json_file='xls2ora.json'

        try:
            with open(json_file,"r+") as f:
                cfg = json.load(f)
        except Exception as e:
            myLog(f"""cfg json file [{json_file}] not valid\n{e}\n""",1)
            print(usage)
            sendicqmsg(1001,f"""cfg json file [{json_file}] not valid\n{e}""")
            return

        try:
            table_in=cfg["table_in"]
            fields_in=cfg["fields_in"]
            
            if arg.find(".json")>-1:
                try:
                    file_in=cfg["file_in"]
                except:
                    err=f'\nNot find key "file_in" in json file'
                    myLog(err,1)
                    print(err)
                    sendicqmsg(1001,err)
                    return
            else:
                file_in=arg
            filename=os.path.basename(file_in)
            
            first_row=cfg.get("first_row",1)
            cols=cfg.get("cols",[])
            format=cfg.get("format","xls")
            truncate=cfg.get("truncate","n")
            separator=cfg.get("separator",",")
            delete=cfg.get("delete","")
        except Exception as e:
            myLog(f"""not valid parameters in xls2ora.json\n""",1)
            print(usage)
            sendicqmsg(1001,f"""not valid parameters in xls2ora.json\n{e}""")
            return
        
        if format not in ['html','xls','csv']:
            myLog("\nformat not valid...\n",1)
            print(usage)
            return

        if format == 'xls' and filename.find('.xlsx')>-1:
            format='xlsx'

        cols_all='N'
        
        if not cols:
            cols_all='Y'
            for i in range(1,len(fields_in.split(","))+1):
                cols.append(i)
        
        myLog(f"Opening {filename} file...\n",1)

        cnt_rows:int=0
        data:list=[]
        
        if format=='html':
            with open(file_in,'r',encoding="utf-8") as f:
                s = f.read()
            soup = bs4.BeautifulSoup(s, 'html.parser')
            myLog("Begin parce...",1)

            trs=soup.find_all('tr')
            cnt_rows=len(trs)
            names=["html_table"]
        elif format=='xls':
            try:
                wb = xlrd.open_workbook(file_in)
            except Exception as e:
                myLog(f"""error open xls file\n{e}\n""",1)
                sendicqmsg(1001,f"""error open xls file\n{e}""")
                return                        
            
            names = wb.sheet_names()
            myLog("The number of worksheets is {0}".format(wb.nsheets),1)
            myLog("Worksheet name(s): {0}".format(names),1)
            ws = wb.sheet_by_index(0)
            cnt_rows = ws.nrows - 1
        elif format=='xlsx':
            try:
                wb = openpyxl.load_workbook(file_in, read_only=True, data_only=True)
            except Exception as e:
                myLog(f"""error open xls file\n{e}\n""",1)
                sendicqmsg(1001,f"""error open xls file\n{e}""")
                return                        
            
            names = wb.sheetnames
            ws = wb[names[0]]

            # myLog("The number of worksheets is {0}".format(wb.nsheets),1)
            myLog("Worksheet name(s): {0}".format(names),1)
            # ws = wb.sheet_by_index(0)
            cnt_rows = ws.max_row

        elif format=="csv":
            data=file2arr(file_in,separator)
            cnt_rows=len(data)
            names=["csv_table"]


        myLog(f"\n{cnt_rows} rows will load from {filename}, sheet: {names[0]}\n",1)

        start_time_main = time.perf_counter()


        myLog(f"Get data from {filename}...",1)


        if format=='xlsx' and cols_all=='Y':
            for r,row in enumerate(ws):
                if r<first_row:
                    continue
                if not any(cell.value for cell in row):
                    break
                data.append([str(cell.value) for cell in row])
                myLog(f"row: {r}",2)                
            # elif format=='xls':
        else:
            for i in range(first_row,cnt_rows+1):
                row=[]
                try:
                    for j in cols:
                        val=''
                        if str(j).find('&filename')>-1:
                            val=filename.replace('.xls','')
                        else:
                            if format=='xls':
                                val=str(ws.cell(i, j-1).value).strip() if ws.cell(i, j-1).value else ''
                            if format=='xlsx':
                                val=str(ws.cell(row=i, column=j).value).strip() if ws.cell(row=i, column=j).value else ''
                            elif format=='html':
                                val=trs[i].find_all("td")[j-1].getText().strip()
                            # elif format=='csv':
                            #     val=csv[i][j-1].strip()
                        # if val:
                        row.append(val)
                    if not row:
                        break
                    data.append(row)
                    myLog(f"row: {i}",2)
                except:
                    pass

       
        if truncate=="Y":
            myLog(f"\ntruncate {table_in}\n",1)
            if truncate_table(table_in=table_in)<0:
                return
        
        if delete:
            if delete.find('&filename')>-1:
                delete=delete.replace('&filename',filename.replace('.xls',''))
            myLog(f"delete {table_in} where {delete}",1)
            if truncate_table(table_in=table_in,delete=delete)<0:
                return            

        myLog(f"send data to {table_in}...",1)
        res=ins_to_ora(data_in=data,table_in=table_in,fields_in=fields_in)
        end = time.perf_counter()

        # total_time = sec2hours(sec)

        msg = f"""{filename}, total rows: [{cnt_rows}], total time: {end-start_time_main:0.7f} s\nresult: {res}"""
        myLog(msg,1)
        sendicqmsg(1001,msg)
        myLog("END")  
        decl_log(tin="", cnt=cnt_rows, decl=nm)  
    except:
        e = traceback.format_exc()
        myLog(e,1)
        sendicqmsg(1001,e)

def main():
    do_xls2ora()

if __name__ == "__main__":
    main()